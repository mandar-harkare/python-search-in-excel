import os
import openpyxl
import questionary


app = str(input("Enter the search keyword: " ))

## Get the files from current directory for auto complete
directory = os.getcwd()
files = os.listdir(directory)

file = questionary.autocomplete(
   'Enter the path of excel file:',
   choices=files).ask()

data = []
columns = []
sr = 1

wb = openpyxl.load_workbook(file, read_only=True)
sheetnames = questionary.checkbox(
    "Select sheets to search in:", choices=wb.sheetnames
).ask()
print("Searching in following sheets:")
print(sheetnames)
for sheet in sheetnames:
    ws = wb[sheet]

    for row in ws.rows:
        if not columns:
            columns = (cell.value for cell in row)
        for cell in row:
            if cell.value and str(cell.value).lower() == app.lower():
                data.append((cell.value for cell in row))

wb2 = openpyxl.Workbook() 
ws1 = wb2.active
# print(data)
ws1.append(columns)
print("Text found in " + str(len(data)) + " rows & copied into " + app+".xlsx")
for page in data:
    ws1.append(page)
    ws1['A'+str(sr + 1)]  = str(sr)
    sr = int(sr) + 1
wb2.save(app+".xlsx")
